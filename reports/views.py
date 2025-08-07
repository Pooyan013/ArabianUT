from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from core.models import RepairJob
from decimal import Decimal
from .forms import DateRangeFilterForm
import openpyxl
from django.db.models import Sum
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from accounting.models import SalarySlip, Expense, Attendance, Employee, Income
from accounting.forms import AttendanceForm, SalarySlipForm, SalaryAdjustmentForm, CloseSlipForm
@login_required
def operational_report_view(request):
    sortable_fields = [
        ('car__brand', 'Car'),
        ('car__model', 'Model'),
        ('car__plate_number', 'Plate'),
        ('status', 'Status'),
        ('car__registered_at', 'Entry Date'),
        ('expert_inspected_at', 'Expert Visit'),
        ('approved_at', 'Approve Date'),
    ]
    jobs = RepairJob.objects.select_related('car')

    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    if start:
        jobs = jobs.filter(car__registered_at__gte=start)
    if end:
        jobs = jobs.filter(car__registered_at__lte=end)

    if request.GET.get('sign'):
        jobs = jobs.filter(sign_confirmed=True)
    if request.GET.get('lpo'):
        jobs = jobs.filter(lpo_confirmed=True)

    sort_field = request.GET.get('sort', 'car__registered_at')
    allowed_fields = [
        'car__brand', 'car__model', 'car__plate_number',
        'status', 'car__registered_at', 'expert_inspected_at', 'approved_at',
        '-car__brand', '-car__model', '-car__plate_number',
        '-status', '-car__registered_at', '-expert_inspected_at', '-approved_at',
    ]
    if sort_field in allowed_fields:
        jobs = jobs.order_by(sort_field)

    context = {
        'form': DateRangeFilterForm(request.GET or None),
        'report_jobs': jobs,
        'sortable_fields': sortable_fields,
        'active_page': 'operational',

    }
    return render(request, 'reports/operational_report.html', context)


@login_required
def financial_report_view(request):
    form = DateRangeFilterForm(request.GET or None)
    jobs = RepairJob.objects.select_related('car').all()

    columns = [
    ('car__brand', 'Car'),
    ('car__model', 'Model'),
    ('car__plate_number', 'Plate'),
    ('status', 'Status'),
    ('approved_at', 'Approve Date'),
    ('deal', 'Deal Amount'),
    ('lpo_confirmed', 'LPO Confirmed'),
    ('sign_confirmed', 'Sign Confirmed'),
    ('vat', 'VAT (5%)'),
    ('total', 'Total Amount'),]

    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        status = form.cleaned_data.get('status')
        lpo = form.cleaned_data.get('lpo')
        sign = form.cleaned_data.get('sign')

        if start_date:
            jobs = jobs.filter(approved_at__gte=start_date)
        if end_date:
            jobs = jobs.filter(approved_at__lte=end_date)
        if status:
            jobs = jobs.filter(status=status)

        if lpo == 'yes':
            jobs = jobs.filter(lpo_confirmed=True)
        elif lpo == 'no':
            jobs = jobs.filter(lpo_confirmed=False)

        if sign == 'yes':
            jobs = jobs.filter(sign_confirmed=True)
        elif sign == 'no':
            jobs = jobs.filter(sign_confirmed=False)

    sort_field = request.GET.get('sort', '-approved_at')  
    allowed_sorts = [
        'car__brand', '-car__brand',
        'car__model', '-car__model',
        'car__plate_number', '-car__plate_number',
        'status', '-status',
        'approved_at', '-approved_at',
        'deal', '-deal',
    ]

    if sort_field in allowed_sorts:
        jobs = jobs.order_by(sort_field)
    else:
        jobs = jobs.order_by('-approved_at')

    jobs_with_deal = [job for job in jobs if job.deal]
    grand_total_deal = sum(job.deal for job in jobs_with_deal)
    grand_total_vat = grand_total_deal * Decimal('0.05')
    grand_total_final = grand_total_deal + grand_total_vat

    for job in jobs:
        if job.deal:
            job.vat = job.deal * Decimal('0.05')
            job.total = job.deal + job.vat
        else:
            job.vat = None
            job.total = None

    context = {
        'form': form,
        'report_jobs': jobs,
        'grand_total_deal': grand_total_deal,
        'grand_total_vat': grand_total_vat,
        'grand_total_final': grand_total_final,
        'active_page': 'financial',
        'current_sort': sort_field,
        'columns': columns,

    }

    return render(request, 'reports/financial_report.html', context)

@login_required
def export_financial_report_excel(request):
    jobs = RepairJob.objects.select_related('car').filter(deal__isnull=False).order_by('-approved_at')
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    lpo = request.GET.get('lpo')
    sign = request.GET.get('sign')

    if start_date:
        jobs = jobs.filter(approved_at__gte=start_date)
    if end_date:
        jobs = jobs.filter(approved_at__lte=end_date)
    if status:
        jobs = jobs.filter(status=status)
    else:
        jobs = jobs.filter(status='archived')
        
    if lpo == 'yes':
        jobs = jobs.filter(lpo_confirmed=True)
    elif lpo == 'no':
        jobs = jobs.filter(lpo_confirmed=False)
        
    if sign == 'yes':
        jobs = jobs.filter(sign_confirmed=True)
    elif sign == 'no':
        jobs = jobs.filter(sign_confirmed=False)
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Operational Report'

    headers = ['Car Make', 'Model', 'Plate', 'Color', 'Status', 'Entry Date', 'Expert Visit', 'Approve Date']
    sheet.append(headers)

    for job in jobs:
        row = [
            job.car.brand, job.car.model, job.car.plate_number, job.car.color,
            job.get_status_display(),
            job.car.registered_at.strftime('%Y-%m-%d') if job.car.registered_at else '',
            job.expert_inspected_at.strftime('%Y-%m-%d') if job.expert_inspected_at else '',
            job.approved_at.strftime('%Y-%m-%d') if job.approved_at else '',
        ]
        sheet.append(row)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="operational_report.xlsx"'
    workbook.save(response)
    return response

from accounting.models import SalarySlip, Attendance 
from .forms import DateRangeFilterForm

@login_required
def payroll_report_view(request):
    form = DateRangeFilterForm(request.GET or None)
    slips = SalarySlip.objects.select_related('employee').all().order_by('-pay_period_end')

    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        if start_date:
            slips = slips.filter(pay_period_end__gte=start_date)
        if end_date:
            slips = slips.filter(pay_period_end__lte=end_date)
            
    for slip in slips:
        slip.absence_count = Attendance.objects.filter(
            employee=slip.employee,
            date__range=(slip.pay_period_start, slip.pay_period_end)
        ).count()
        
        slip.personal_expenses = Expense.objects.filter(
            employee=slip.employee,
            expense_type='personal',
            transaction_date__range=(slip.pay_period_start, slip.pay_period_end)
        )

    total_cost_sum = sum(slip.cost for slip in slips)

    context = {
        'form': form,
        'salary_slips': slips,
        'total_cost_sum': total_cost_sum,
        'active_page': 'payroll',
    }
    return render(request, 'reports/payroll_report.html', context)

@login_required
def export_operational_excel(request):
    jobs = RepairJob.objects.select_related('car').all()

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sign = request.GET.get('sign')
    lpo = request.GET.get('lpo')
    deal = request.GET.get('deal')

    if start_date:
        jobs = jobs.filter(approved_at__gte=start_date)
    if end_date:
        jobs = jobs.filter(approved_at__lte=end_date)

    if sign == 'on':
        jobs = jobs.filter(sign_confirmed=True)
    if lpo == 'on':
        jobs = jobs.filter(lpo_confirmed=True)
    if deal == 'on':
        jobs = jobs.exclude(deal__isnull=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Operational Report'

    headers = [
        'Car Brand', 'Model', 'Plate', 'Color',
        'Status', 'Entry Date', 'Expert Visit', 'Approve Date'
    ]
    sheet.append(headers)

    for job in jobs:
        row = [
            job.car.brand if job.car else '',
            job.car.model if job.car else '',
            job.car.plate_number if job.car else '',
            job.car.color if job.car else '',
            job.get_status_display(),
            job.car.registered_at.strftime('%Y-%m-%d') if job.car and job.car.registered_at else '',
            job.expert_inspected_at.strftime('%Y-%m-%d') if job.expert_inspected_at else '',
            job.approved_at.strftime('%Y-%m-%d') if job.approved_at else '',
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="operational_report.xlsx"'
    workbook.save(response)
    return response

def payroll_dashboard_view(request):
    if request.method == 'POST':
        if 'add_absence' in request.POST:
            attendance_form = AttendanceForm(request.POST)
            if attendance_form.is_valid():
                attendance_form.save()
                messages.success(request, 'Absence record saved successfully.')
            else:
                messages.error(request, 'Please correct the errors in the attendance form.')
            return redirect('reports:dashboard')

        elif 'create_slip' in request.POST:
            salary_slip_form = SalarySlipForm(request.POST)
            if salary_slip_form.is_valid():
                slip = salary_slip_form.save(commit=False)
                
                last_slip = SalarySlip.objects.filter(employee=slip.employee).order_by('-pay_period_end').first()
                if last_slip:
                    slip.remaining_before = last_slip.rest
                
                slip.save()
                messages.success(request, 'Salary slip created successfully.')
            else:
                messages.error(request, 'Please correct the errors in the salary slip form.')
            return redirect('reports:payroll_dashboard')

    attendance_form = AttendanceForm()
    salary_slip_form = SalarySlipForm()
    recent_absences = Attendance.objects.all()[:10]  
    salary_slips = SalarySlip.objects.select_related('employee').all()

    context = {
        'attendance_form': attendance_form,
        'salary_slip_form': salary_slip_form,
        'recent_absences': recent_absences,
        'salary_slips': salary_slips,
    }
    return render(request, 'reports/dashboard.html', context)


@login_required
def employee_payroll_detail_view(request, employee_id):
    """
    Displays the detailed payroll page for a single employee.
    """
    employee = get_object_or_404(Employee, id=employee_id)
    slips = SalarySlip.objects.filter(employee=employee).order_by('-pay_period_start')
    absences = Attendance.objects.filter(employee=employee)
    
    # فرم‌ها برای نمایش در صفحه
    adjustment_form = SalaryAdjustmentForm()
    close_slip_form = CloseSlipForm()

    context = {
        'employee': employee,
        'slips': slips,
        'absences': absences,
        'adjustment_form': adjustment_form,
        'close_slip_form': close_slip_form,
    }
    return render(request, 'reports/employee_detail.html', context)

@login_required
def add_salary_adjustment_view(request, slip_id):
    """
    Handles the submission of the salary adjustment form.
    """
    if request.method == 'POST':
        slip = get_object_or_404(SalarySlip, id=slip_id)
        form = SalaryAdjustmentForm(request.POST)

        if form.is_valid():
            cd = form.cleaned_data
            
            # مقادیر جدید به مقادیر قبلی اضافه می‌شوند
            if cd.get('paid_add'): slip.paid += cd['paid_add']
            if cd.get('extra_h_add'): slip.extra_h += cd['extra_h_add']
            if cd.get('mines_h_add'): slip.mines_h += cd['mines_h_add']
            if cd.get('extra_add'): slip.extra += cd['extra_add']
            if cd.get('mines_add'): slip.mines += cd['mines_add']
            
            new_desc = cd.get('description')
            if new_desc:
                slip.description = (slip.description or "") + f"\n- Adjustment: {new_desc}"

            slip.save()
            messages.success(request, "Adjustments saved successfully.")
        else:
            messages.error(request, "Invalid data submitted.")
            
    return redirect('reports:employee_detail', employee_id=slip.employee.id)

@login_required
def close_salary_slip_view(request, slip_id):
    """
    Closes the current salary slip and transfers the remaining balance.
    """
    if request.method == 'POST':
        slip_to_close = get_object_or_404(SalarySlip, id=slip_id)
        form = CloseSlipForm(request.POST)

        if form.is_valid():
            slip_to_close.is_closed = True
            slip_to_close.save()

            next_period_start = slip_to_close.pay_period_end + timezone.timedelta(days=1)
            next_slip, created = SalarySlip.objects.get_or_create(
                employee=slip_to_close.employee,
                pay_period_start=next_period_start,
                defaults={
                    'pay_period_end': next_period_start + timezone.timedelta(days=29), # فرض دوره یک ماهه
                    'remaining_before': slip_to_close.rest, 
                }
            )
            if not created:
                next_slip.remaining_before += slip_to_close.rest
                next_slip.save()

            messages.success(request, f"Period for {slip_to_close.employee.full_name} closed. Remaining balance transferred.")
        else:
            messages.error(request, "You must confirm to close the period.")

    return redirect('reports:employee_detail', employee_id=slip_to_close.employee.id)


@login_required
def profit_report_view(request):
    from_date = request.GET.get('from')
    to_date = request.GET.get('to')

    incomes = Income.objects.all()
    expenses = Expense.objects.all()

    if from_date:
        incomes = incomes.filter(transaction_date__date__gte=from_date)
        expenses = expenses.filter(transaction_date__date__gte=from_date)
    if to_date:
        incomes = incomes.filter(transaction_date__date__lte=to_date)
        expenses = expenses.filter(transaction_date__date__lte=to_date)

    total_income = incomes.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_expense = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expense = abs(total_expense)

    net_profit = total_income - total_expense

    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'incomes': incomes.order_by('-transaction_date'),
        'expenses': expenses.order_by('-transaction_date'),
        'from_date': from_date, 
        'to_date': to_date,
        'active_page': 'profit-report',
    }
    return render(request, 'reports/profit_report.html', context)